#!/usr/bin/env -S uv run --script
# -*- coding: utf-8 -*-
# /// script
# dependencies = [
#   "numpy>=1.21",
#   "openpyxl>=3.0",
# ]
# ///
"""
Joint Stiffness & Damping 回归计算 + 格式化 XLSX 输出
======================================================
基于阻抗模型: tau = K·(theta_ref − theta_sta) + D·(theta_dot_ref − theta_dot_sta)

用法:
  uv run compute.py                                                            # 本地运行
  uv run --script https://cdn.jsdelivr.net/gh/...@main/compute.py              # 远程运行 (uv ≥ 0.4)
  curl -sL https://cdn.jsdelivr.net/gh/...@main/compute.py | uv run -         # 管道运行 (Unix)
  (Invoke-WebRequest https://cdn.jsdelivr.net/gh/...@main/compute.py).Content | uv run -  # 管道运行 (Win)

关节分组 (29 -> 7-7-3-6-6, 27 -> 7-7-1-6-6):
  7-7 和 6-6 并排对比; 联合适的数量和分组自动适配。
支持多组 CSV -> 多个 Sheet。
"""

import argparse, glob, os, sys, re
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# 0. 样式定义
# ============================================================
HEADER_FILL   = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT   = Font(name="Microsoft YaHei", bold=True, color="FFFFFF", size=11)
TITLE_FONT    = Font(name="Microsoft YaHei", bold=True, size=14, color="1F4E79")
GROUP_FONT    = Font(name="Microsoft YaHei", bold=True, size=11, color="2E75B6")
NORMAL_FONT   = Font(name="Consolas", size=10)
WARN_FONT     = Font(name="Consolas", size=10, color="999999")
THIN_BORDER   = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
CENTER_ALIGN  = Alignment(horizontal="center", vertical="center")
LEFT_ALIGN    = Alignment(horizontal="left", vertical="center")
LIGHT_FILL    = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
DARK_FILL     = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

# 异常程度颜色 — 用字体颜色标注, 不改底纹
GOOD_COLOR    = "006100"   # 深绿: 正常 / 较好
WEAK_COLOR    = "9C6500"   # 棕色: 一般 / 偏弱
BAD_COLOR     = "C00000"   # 深红: 异常 / 较差


def get_quality_font(value, metric, valid_scores):
    """
    返回异常标注用的 Font 对象 (只改颜色，不改底纹)。
    metric: "R2" / "RMSE" / "K" / "D"
    """
    if valid_scores is None:
        return None
    if metric == "R2":
        if value < 0.6:   return Font(name="Consolas", size=10, color=BAD_COLOR, bold=True)
        if value < 0.8:   return Font(name="Consolas", size=10, color=WEAK_COLOR, bold=True)
        return Font(name="Consolas", size=10, color=GOOD_COLOR)
    elif metric == "RMSE":
        arr = valid_scores.get("rmse_vals")
        if arr is None or len(arr) == 0:
            return None
        p75 = np.percentile(arr, 75)
        p90 = np.percentile(arr, 90)
        if value > p90:   return Font(name="Consolas", size=10, color=BAD_COLOR, bold=True)
        if value > p75:   return Font(name="Consolas", size=10, color=WEAK_COLOR, bold=True)
        return Font(name="Consolas", size=10, color=GOOD_COLOR)
    elif metric == "K":
        arr = valid_scores.get("k_vals")
        if arr is None or len(arr) == 0: return None
        mean, std = np.mean(arr), np.std(arr)
        if std < 1e-12: return None
        z = abs(value - mean) / std
        if z > 2.0:     return Font(name="Consolas", size=10, color=BAD_COLOR, bold=True)
        if z > 1.0:     return Font(name="Consolas", size=10, color=WEAK_COLOR, bold=True)
        return Font(name="Consolas", size=10, color=GOOD_COLOR)
    elif metric == "D":
        arr = valid_scores.get("d_vals")
        if arr is None or len(arr) == 0: return None
        mean, std = np.mean(arr), np.std(arr)
        if std < 1e-12: return None
        z = abs(value - mean) / std
        if z > 2.0:     return Font(name="Consolas", size=10, color=BAD_COLOR, bold=True)
        if z > 1.0:     return Font(name="Consolas", size=10, color=WEAK_COLOR, bold=True)
        return Font(name="Consolas", size=10, color=GOOD_COLOR)
    return None


def style_header_row(ws, row, col_start, col_end):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER


def style_data_cell(ws, row, col, value, fmt="0.0000", is_warn=False, fill=None, font=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font if font is not None else (WARN_FONT if is_warn else NORMAL_FONT)
    cell.alignment = CENTER_ALIGN
    cell.border = THIN_BORDER
    if fill is not None:
        cell.fill = fill
    if isinstance(value, (int, float)) and not is_warn and fmt:
        cell.number_format = fmt
    return cell


def apply_banding(ws, start_row, end_row, col_start, col_end):
    """每个子表格独立: 第1行浅色, 第2行深色, 交替重复"""
    for r in range(start_row, end_row + 1):
        fill = DARK_FILL if (r - start_row) % 2 == 1 else LIGHT_FILL
        for c in range(col_start, col_end + 1):
            ws.cell(row=r, column=c).fill = fill


# ============================================================
# 1. 查找 CSV 配对
# ============================================================
def find_csv_pairs(directory):
    ref_files = sorted(glob.glob(os.path.join(directory, "dof_ref_*.csv")))
    pairs = []
    for ref in ref_files:
        base = os.path.basename(ref)
        m = re.match(r"dof_ref_(.+)\.csv", base)
        if not m:
            continue
        suffix = m.group(1)
        sta = os.path.join(directory, f"dof_sta_{suffix}.csv")
        if os.path.exists(sta):
            label = suffix.replace("_", " ").strip()
            pairs.append((label, ref, sta))
    return pairs


# ============================================================
# 2. 数据处理
# ============================================================
def _load_csv_safe(path):
    """Load CSV with numpy, skipping lines that are corrupted (null bytes, wrong column count)."""
    # First pass: read raw lines, skip obviously bad ones
    import csv
    good_lines = []
    expected_cols = None
    with open(path, "r", encoding="utf-8", errors="replace") as f:
        reader = csv.reader(f)
        for row in reader:
            if expected_cols is None:
                # header
                expected_cols = len(row)
                good_lines.append(",".join(row))
                continue
            # skip empty lines or lines with null bytes
            if not row or all(c == "\0" or c == "" for c in row):
                continue
            if len(row) != expected_cols:
                continue
            # replace null bytes in any field
            cleaned = [c.replace("\0", "") for c in row]
            good_lines.append(",".join(cleaned))
    import tempfile
    fd, tmp = tempfile.mkstemp(suffix=".csv")
    os.close(fd)
    try:
        with open(tmp, "w", encoding="utf-8") as f:
            f.write("\n".join(good_lines))
        return np.genfromtxt(tmp, delimiter=",", names=True, dtype=None, encoding="utf-8")
    finally:
        os.unlink(tmp)


def load_and_compute(ref_path, sta_path):
    ref_data = _load_csv_safe(ref_path)
    sta_data = _load_csv_safe(sta_path)

    ref_cols = ref_data.dtype.names
    sta_cols = sta_data.dtype.names

    n_joints = sum(1 for c in ref_cols if c.startswith("joint_pos"))
    ref_ts = ref_data["timestamp"]
    sta_ts = sta_data["timestamp"]

    mask = (sta_ts >= ref_ts.min()) & (sta_ts <= ref_ts.max())
    sta_ts_aligned = sta_ts[mask]

    results = []
    for j in range(n_joints):
        pos_ref_col = f"joint_pos{j}"
        vel_ref_col = f"joint_vel{j}"
        pos_sta_col = f"joint_pos{j}"
        vel_sta_col = f"joint_vel{j}"
        tor_sta_col = f"joint_tor{j}"

        if pos_ref_col not in ref_cols or pos_sta_col not in sta_cols:
            results.append({"joint": j, "K": np.nan, "D": np.nan, "R2": np.nan,
                            "RMSE": np.nan, "MAE": np.nan, "ExplVar": np.nan,
                            "dp_std": np.nan, "tq_std": np.nan, "valid": False})
            continue

        pos_ref_interp = np.interp(sta_ts_aligned, ref_ts, ref_data[pos_ref_col])
        vel_ref_interp = np.interp(sta_ts_aligned, ref_ts, ref_data[vel_ref_col])
        pos_sta = sta_data[pos_sta_col][mask]
        vel_sta = sta_data[vel_sta_col][mask]
        tor_sta = sta_data[tor_sta_col][mask]

        delta_pos = pos_ref_interp - pos_sta
        delta_vel = vel_ref_interp - vel_sta
        tq_std = float(np.std(tor_sta))

        if tq_std < 1e-12:
            results.append({"joint": j, "K": 0.0, "D": 0.0, "R2": np.nan,
                            "RMSE": np.nan, "MAE": np.nan, "ExplVar": np.nan,
                            "dp_std": float(np.std(delta_pos)), "tq_std": 0.0, "valid": False})
            continue

        X = np.column_stack([delta_pos, delta_vel])
        y = tor_sta

        try:
            beta, _, _, _ = np.linalg.lstsq(X, y, rcond=None)
            K, D = float(beta[0]), float(beta[1])
            y_pred = X @ beta
            residuals_vec = y - y_pred
            n = len(y)
            ss_res = np.sum(residuals_vec ** 2)
            ss_tot = np.sum((y - np.mean(y)) ** 2)
            r2 = float(1.0 - ss_res / ss_tot) if ss_tot > 1e-15 else np.nan
            rmse = float(np.sqrt(ss_res / n))
            mae = float(np.mean(np.abs(residuals_vec)))
            ev = float(1.0 - np.var(residuals_vec) / np.var(y)) if np.var(y) > 1e-15 else np.nan
            results.append({"joint": j, "K": K, "D": D,
                            "R2": r2, "RMSE": rmse, "MAE": mae, "ExplVar": ev,
                            "dp_std": float(np.std(delta_pos)), "tq_std": tq_std, "valid": True})
        except np.linalg.LinAlgError:
            results.append({"joint": j, "K": np.nan, "D": np.nan,
                            "R2": np.nan, "RMSE": np.nan, "MAE": np.nan, "ExplVar": np.nan,
                            "dp_std": float(np.std(delta_pos)), "tq_std": tq_std, "valid": False})

    return results, n_joints, len(ref_ts), len(sta_ts)


# ============================================================
# 3. 分组逻辑
# ============================================================
def get_groups(n_joints):
    if n_joints == 29:
        return [
            ("A", list(range(0, 7))),
            ("B", list(range(7, 14))),
            ("C", list(range(14, 17))),
            ("D", list(range(17, 23))),
            ("E", list(range(23, 29))),
        ]
    elif n_joints == 27:
        return [
            ("A", list(range(0, 7))),
            ("B", list(range(7, 14))),
            ("C", list(range(14, 15))),
            ("D", list(range(15, 21))),
            ("E", list(range(21, 27))),
        ]
    else:
        g1_end = min(7, n_joints)
        g2_end = min(14, n_joints)
        rem = n_joints - 14
        if rem <= 0:
            return [("A", list(range(0, g1_end))), ("B", list(range(g1_end, n_joints)))]
        d_size = min(6, rem)
        e_size = min(6, rem - d_size)
        c_size = rem - d_size - e_size
        if c_size < 0:
            c_size = 0
        c_start = min(g2_end, n_joints)
        c_end = c_start + c_size
        d_start = c_end
        d_end = d_start + d_size
        e_start = d_end
        e_end = e_start + e_size
        groups = []
        if g1_end > 0: groups.append(("A", list(range(0, g1_end))))
        if g2_end > g1_end: groups.append(("B", list(range(g1_end, g2_end))))
        if c_end > c_start: groups.append(("C", list(range(c_start, c_end))))
        if d_end > d_start: groups.append(("D", list(range(d_start, d_end))))
        if e_end > e_start: groups.append(("E", list(range(e_start, e_end))))
        return groups


def fmt_metric(val, valid, nan_text="—"):
    """非空数值 → float, 否则返回 nan_text"""
    if valid and not np.isnan(val):
        return float(val)
    return nan_text


# 指标说明文本
METRICS_LEGEND = [
    ("R² (决定系数)", "Coefficient of Determination",
     "衡量回归模型对观测值的拟合程度。取值范围 [0, 1]，越接近 1 表示模型解释力越强。"
     "R² = 1 - SS_res / SS_tot"),
    ("RMSE (均方根误差)", "Root Mean Squared Error",
     "预测值与真实值偏差的平方和与样本数比值的平方根。与扭矩同单位 (Nm)，越小越好。"
     "RMSE = sqrt(Σ(y_true - y_pred)² / n)"),
    ("MAE (平均绝对误差)", "Mean Absolute Error",
     "预测值与真实值偏差绝对值的平均值。与扭矩同单位 (Nm)，对异常值不如 RMSE 敏感。"
     "MAE = Σ|y_true - y_pred| / n"),
    ("ExplVar (可解释方差)", "Explained Variance Score",
     "衡量模型捕捉数据变异程度的能力。取值范围 (-∞, 1]，越接近 1 越好。"
     "ExplVar = 1 - Var(y_true - y_pred) / Var(y_true)"),
    ("K (刚度)", "Stiffness",
     "位置误差到力矩的映射系数。单位: Nm/rad。正值表示弹性恢复力。"),
    ("D (阻尼)", "Damping",
     "速度误差到力矩的映射系数。单位: Nm·s/rad。正值表示粘性阻尼。"),
]


# ============================================================
# 4. 写入数据 Sheet（每个 CSV 配对一张，只含数据表）
# ============================================================
def write_data_sheet(ws, results, n_joints, label, n_ref, n_sta, valid_scores):
    groups = get_groups(n_joints)

    # 列宽预设 (数据表 B~M 统一 13)
    ws.column_dimensions["A"].width = 2
    for col_letter in ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M"]:
        ws.column_dimensions[col_letter].width = 13

    row = 1

    # -------- 标题 --------
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=13)
    title_cell = ws.cell(row=row, column=2,
                         value=f"Joint Stiffness & Damping   [{label}]")
    title_cell.font = TITLE_FONT
    title_cell.alignment = LEFT_ALIGN
    row += 1

    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=13)
    info_cell = ws.cell(row=row, column=2,
                        value=f"Joints: {n_joints}  |  分组: {'-'.join(str(len(g[1])) for g in groups)}"
                              f"  |  Ref: {n_ref} 行  |  Sta: {n_sta} 行")
    info_cell.font = Font(name="Microsoft YaHei", size=9, color="666666")
    row += 2

    # -------- 整理 side-by-side 对 & standalone --------
    side_by_side = []
    standalone = []
    used = set()
    for g in groups:
        if g[0] in used:
            continue
        if g[0] == "A" and any(gg[0] == "B" for gg in groups):
            b_indices = [gg for gg in groups if gg[0] == "B"]
            if b_indices:
                side_by_side.append((g, b_indices[0], "7-7"))
                used.update(["A", "B"])
        elif g[0] == "D" and any(gg[0] == "E" for gg in groups):
            e_indices = [gg for gg in groups if gg[0] == "E"]
            if e_indices:
                side_by_side.append((g, e_indices[0], "6-6"))
                used.update(["D", "E"])
        else:
            standalone.append(g)
            used.add(g[0])

    # -------- 构建异常程度评分基准 --------
    valid_results = [r for r in results if r["valid"]]
    if valid_scores is None and valid_results:
        valid_scores = {
            "r2_vals":   np.array([r["R2"] for r in valid_results if not np.isnan(r["R2"])]),
            "rmse_vals": np.array([r["RMSE"] for r in valid_results if not np.isnan(r["RMSE"])]),
            "k_vals":    np.array([r["K"] for r in valid_results]),
            "d_vals":    np.array([r["D"] for r in valid_results]),
        }

    # -------- 写 side-by-side 块 (新增 RMSE 列 + 异常着色)--------
    for left_g, right_g, tag in side_by_side:
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=13)
        ws.cell(row=row, column=2,
                value=f"{tag} 并排对比   "
                      f"({left_g[0]}组: J{left_g[1][0]}~J{left_g[1][-1]}   |   "
                      f"{right_g[0]}组: J{right_g[1][0]}~J{right_g[1][-1]})").font = GROUP_FONT
        row += 1

        # 表头
        sb_headers = ["J#", "K", "D", "R²", "RMSE", "", "J#", "K", "D", "R²", "RMSE"]
        for ci, h in enumerate(sb_headers):
            if h:
                ws.cell(row=row, column=2 + ci, value=h)
        style_header_row(ws, row, 2, 12)
        sep_cell = ws.cell(row=row, column=7)
        sep_cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        row += 1

        n_rows = max(len(left_g[1]), len(right_g[1]))
        data_start = row
        for i in range(n_rows):
            band_fill = DARK_FILL if i % 2 == 1 else LIGHT_FILL

            # 左侧 (col 2-6)
            if i < len(left_g[1]):
                j = left_g[1][i]; r = results[j]
                kf = get_quality_font(r["K"], "K", valid_scores) if r["valid"] else None
                df = get_quality_font(r["D"], "D", valid_scores) if r["valid"] else None
                rf = get_quality_font(r["R2"], "R2", valid_scores) if r["valid"] else None
                mf = get_quality_font(r["RMSE"], "RMSE", valid_scores) if r["valid"] else None
                style_data_cell(ws, row, 2, j, fmt="0", fill=band_fill)
                style_data_cell(ws, row, 3, r["K"],
                                is_warn=not r["valid"], fill=band_fill, font=kf)
                style_data_cell(ws, row, 4, r["D"],
                                is_warn=not r["valid"], fill=band_fill, font=df)
                style_data_cell(ws, row, 5, fmt_metric(r["R2"], r["valid"]),
                                fmt="0.0000" if r["valid"] else "", fill=band_fill, font=rf)
                style_data_cell(ws, row, 6, fmt_metric(r["RMSE"], r["valid"]),
                                fmt="0.0000" if r["valid"] else "",
                                is_warn=not r.get("RMSE") or np.isnan(r.get("RMSE", np.nan)),
                                fill=band_fill, font=mf)

            # 分隔列
            ws.cell(row=row, column=7).border = THIN_BORDER
            ws.cell(row=row, column=7).fill = PatternFill(
                start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

            # 右侧 (col 8-12)
            if i < len(right_g[1]):
                j = right_g[1][i]; r = results[j]
                kf = get_quality_font(r["K"], "K", valid_scores) if r["valid"] else None
                df = get_quality_font(r["D"], "D", valid_scores) if r["valid"] else None
                rf = get_quality_font(r["R2"], "R2", valid_scores) if r["valid"] else None
                mf = get_quality_font(r["RMSE"], "RMSE", valid_scores) if r["valid"] else None
                style_data_cell(ws, row, 8, j, fmt="0", fill=band_fill)
                style_data_cell(ws, row, 9, r["K"],
                                is_warn=not r["valid"], fill=band_fill, font=kf)
                style_data_cell(ws, row, 10, r["D"],
                                is_warn=not r["valid"], fill=band_fill, font=df)
                style_data_cell(ws, row, 11, fmt_metric(r["R2"], r["valid"]),
                                fmt="0.0000" if r["valid"] else "", fill=band_fill, font=rf)
                style_data_cell(ws, row, 12, fmt_metric(r["RMSE"], r["valid"]),
                                fmt="0.0000" if r["valid"] else "",
                                is_warn=not r.get("RMSE") or np.isnan(r.get("RMSE", np.nan)),
                                fill=band_fill, font=mf)

            row += 1

        # 不需要单独 apply_banding，已在循环内处理
        row += 1

    # -------- 写 standalone 块 (新增 RMSE 列 + 异常着色)--------
    for g in standalone:
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
        ws.cell(row=row, column=2,
                value=f"{g[0]}组: J{g[1][0]}~J{g[1][-1]}  ({len(g[1])} joints)").font = GROUP_FONT
        row += 1

        for ci, h in enumerate(["J#", "K", "D", "R²", "RMSE"]):
            ws.cell(row=row, column=2 + ci, value=h)
        style_header_row(ws, row, 2, 6)
        row += 1

        data_start = row
        for i_rel, j in enumerate(g[1]):
            r = results[j]
            band_fill = DARK_FILL if i_rel % 2 == 1 else LIGHT_FILL
            kf = get_quality_font(r["K"], "K", valid_scores) if r["valid"] else None
            df = get_quality_font(r["D"], "D", valid_scores) if r["valid"] else None
            rf = get_quality_font(r["R2"], "R2", valid_scores) if r["valid"] else None
            mf = get_quality_font(r["RMSE"], "RMSE", valid_scores) if r["valid"] else None
            style_data_cell(ws, row, 2, j, fmt="0", fill=band_fill)
            style_data_cell(ws, row, 3, r["K"],
                            is_warn=not r["valid"], fill=band_fill, font=kf)
            style_data_cell(ws, row, 4, r["D"],
                            is_warn=not r["valid"], fill=band_fill, font=df)
            style_data_cell(ws, row, 5, fmt_metric(r["R2"], r["valid"]),
                            fmt="0.0000" if r["valid"] else "", fill=band_fill, font=rf)
            style_data_cell(ws, row, 6, fmt_metric(r["RMSE"], r["valid"]),
                            fmt="0.0000" if r["valid"] else "",
                            is_warn=not r.get("RMSE") or np.isnan(r.get("RMSE", np.nan)),
                            fill=band_fill, font=mf)
            row += 1
        row += 1

    ws.freeze_panes = "A1"


# ============================================================
# 4b. 写入汇总 Sheet（首表：Summary + Legend，不干扰数据表列宽）
# ============================================================
def write_summary_sheet(ws, all_data):
    """
    all_data: list of (label, results, n_joints, n_ref, n_sta)
    """
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 26
    ws.column_dimensions["D"].width = 38
    ws.column_dimensions["E"].width = 26

    row = 1

    # -------- 标题 --------
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
    ws.cell(row=row, column=2, value="Joint Stiffness & Damping — Summary & Legend").font = TITLE_FONT
    row += 2

    # -------- 每组 Summary --------
    for label, results, n_joints, n_ref, n_sta in all_data:
        groups = get_groups(n_joints)
        valid = [r for r in results if r["valid"]]
        invalid = [r for r in results if not r["valid"]]

        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        ws.cell(row=row, column=2,
                value=f"Summary   [{label}]   Joints: {n_joints}  "
                      f"分组: {'-'.join(str(len(g[1])) for g in groups)}  "
                      f"Ref: {n_ref}  Sta: {n_sta}").font = GROUP_FONT
        row += 1

        stats_headers = ["Metric", "Mean ± Std", "Range (Min / Max)", "说明"]
        for ci, h in enumerate(stats_headers):
            ws.cell(row=row, column=2 + ci, value=h)
        style_header_row(ws, row, 2, 5)
        row += 1

        if valid:
            Ks  = np.array([r["K"] for r in valid])
            Ds  = np.array([r["D"] for r in valid])
            R2s = np.array([r["R2"] for r in valid if not np.isnan(r["R2"])])
            RMSEs = np.array([r["RMSE"] for r in valid if not np.isnan(r["RMSE"])])
            MAEs  = np.array([r["MAE"] for r in valid if not np.isnan(r["MAE"])])

            stat_rows = [
                ("Valid Joints",
                 f"{len(valid)} / {n_joints}",
                 "",
                 f"无激励(τ≡0)关节: {len(invalid)} 个"),
                ("Stiffness K (Nm/rad)",
                 f"{Ks.mean():.4f} ± {Ks.std():.4f}",
                 f"[{Ks.min():.4f}, {Ks.max():.4f}]",
                 "位置误差→力矩映射系数"),
                ("Damping D (Nm·s/rad)",
                 f"{Ds.mean():.4f} ± {Ds.std():.4f}",
                 f"[{Ds.min():.4f}, {Ds.max():.4f}]",
                 "速度误差→力矩映射系数"),
            ]
            if len(R2s) > 0:
                stat_rows.append(("R² (决定系数)",
                                  f"{R2s.mean():.4f} ± {R2s.std():.4f}",
                                  f"[{R2s.min():.4f}, {R2s.max():.4f}]",
                                  "1=完美拟合, 0=无解释力"))
            if len(RMSEs) > 0:
                stat_rows.append(("RMSE (均方根误差, Nm)",
                                  f"{RMSEs.mean():.4f} ± {RMSEs.std():.4f}",
                                  f"[{RMSEs.min():.4f}, {RMSEs.max():.4f}]",
                                  "越小越好, 对异常值敏感"))
            if len(MAEs) > 0:
                stat_rows.append(("MAE (平均绝对误差, Nm)",
                                  f"{MAEs.mean():.4f} ± {MAEs.std():.4f}",
                                  f"[{MAEs.min():.4f}, {MAEs.max():.4f}]",
                                  "越小越好, 对异常值不敏感"))

            stat_start = row
            for sr in stat_rows:
                for ci, v in enumerate(sr):
                    cell = ws.cell(row=row, column=2 + ci, value=v)
                    cell.font = Font(name="Consolas", bold=(ci == 0), size=10)
                    cell.alignment = CENTER_ALIGN if ci > 0 else LEFT_ALIGN
                    cell.border = THIN_BORDER
                ws.row_dimensions[row].height = 22
                row += 1
            apply_banding(ws, stat_start, row - 1, 2, 5)
        else:
            ws.cell(row=row, column=2, value="No valid joints").font = WARN_FONT
            row += 1

        row += 1  # 组间空行

    # -------- 指标说明 Legend（全局仅一份）--------
    row += 1
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    ws.cell(row=row, column=2, value="Metrics Legend").font = GROUP_FONT
    row += 1

    legend_headers = ["指标", "英文名", "含义 & 公式"]
    for ci, h in enumerate(legend_headers):
        ws.cell(row=row, column=2 + ci, value=h)
    style_header_row(ws, row, 2, 4)
    row += 1

    LEGEND_FILL_LIGHT = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    LEGEND_FILL_DARK  = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    for li, (name_cn, name_en, desc) in enumerate(METRICS_LEGEND):
        fill = LEGEND_FILL_DARK if li % 2 == 1 else LEGEND_FILL_LIGHT
        vals = [name_cn, name_en, desc]
        for ci, v in enumerate(vals):
            cell = ws.cell(row=row, column=2 + ci, value=v)
            cell.font = Font(name="Microsoft YaHei", bold=(ci == 0), size=9,
                             color="333333")
            cell.alignment = LEFT_ALIGN if ci == 2 else CENTER_ALIGN
            cell.border = THIN_BORDER
            cell.fill = fill
        desc_cell = ws.cell(row=row, column=4)
        desc_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = 50
        row += 1

    # 字体颜色说明
    row += 1
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    ws.cell(row=row, column=2, value="字体颜色说明").font = Font(
        name="Microsoft YaHei", bold=True, size=9, color="333333")
    row += 1
    color_legend = [
        ("深绿 006100", GOOD_COLOR,  "正常 / 较好"),
        ("棕色 9C6500", WEAK_COLOR,  "一般 / 偏弱 (R²<0.8, z-score 1~2, 或 RMSE>P75)"),
        ("深红 C00000", BAD_COLOR,   "异常 / 较差 (R²<0.6, z-score>2, 或 RMSE>P90)"),
        ("灰色 999999", "999999",   "无效数据 (力矩恒为零)"),
    ]
    for cl_name, cl_hex, cl_desc in color_legend:
        cell_b = ws.cell(row=row, column=2, value=cl_name)
        cell_b.font = Font(name="Consolas", size=10, color=cl_hex, bold=True)
        cell_b.alignment = CENTER_ALIGN
        cell_b.border = THIN_BORDER

        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
        cell_c = ws.cell(row=row, column=3, value=cl_desc)
        cell_c.font = Font(name="Microsoft YaHei", size=9, color="666666")
        cell_c.alignment = LEFT_ALIGN
        cell_c.border = THIN_BORDER
        ws.row_dimensions[row].height = 18
        row += 1

    ws.freeze_panes = "A1"


# ============================================================
# 5. Main
# ============================================================
def main():
    parser = argparse.ArgumentParser(
        description="Joint Stiffness & Damping 回归计算",
        epilog="示例: uv run compute.py -C ./data -o ./results",
    )
    parser.add_argument("-C", "--cwd", default=".",
                        help="CSV 输入文件所在目录 (默认: 当前目录)")
    parser.add_argument("-o", "--output-dir", default=None,
                        help="XLSX 输出目录 (默认: 与输入目录相同)")
    args = parser.parse_args()

    work_dir = os.path.abspath(args.cwd)
    out_dir = os.path.abspath(args.output_dir) if args.output_dir else work_dir

    pairs = find_csv_pairs(work_dir)

    if not pairs:
        print(f"ERROR: No dof_ref_*/dof_sta_* pairs found in '{work_dir}'")
        sys.exit(1)

    for label, ref_path, sta_path in pairs:
        # 提取时间后缀用于文件名
        base = os.path.basename(ref_path)
        m = re.match(r"dof_ref_(.+)\.csv", base)
        suffix = m.group(1) if m else label.replace(" ", "_")

        print(f"\nProcessing: {label}")
        print(f"  Ref: {os.path.basename(ref_path)}")
        print(f"  Sta: {os.path.basename(sta_path)}")

        results, n_joints, n_ref, n_sta = load_and_compute(ref_path, sta_path)
        groups = get_groups(n_joints)
        valid = [r for r in results if r["valid"]]
        print(f"  Joints: {n_joints}  Groups: {'-'.join(str(len(g[1])) for g in groups)}  "
              f"Valid: {len(valid)}")

        # 构建 valid_scores
        cross_scores = None
        if valid:
            cross_scores = {
                "r2_vals":   np.array([r["R2"] for r in valid if not np.isnan(r["R2"])]),
                "rmse_vals": np.array([r["RMSE"] for r in valid if not np.isnan(r["RMSE"])]),
                "k_vals":    np.array([r["K"] for r in valid]),
                "d_vals":    np.array([r["D"] for r in valid]),
            }

        wb = Workbook()
        wb.remove(wb.active)

        # Sheet 1: Summary + Legend
        ws_summary = wb.create_sheet(title="Summary")
        write_summary_sheet(ws_summary, [(label, results, n_joints, n_ref, n_sta)])

        # Sheet 2: 数据表
        ws = wb.create_sheet(title=label.replace(" ", "_")[:31])
        write_data_sheet(ws, results, n_joints, label, n_ref, n_sta, cross_scores)

        output_path = os.path.join(out_dir, f"stiffness_damping_{suffix}.xlsx")
        os.makedirs(out_dir, exist_ok=True)
        wb.save(output_path)
        print(f"  Saved: {os.path.basename(output_path)}")

    print(f"\nDone. {len(pairs)} file(s) generated.")


if __name__ == "__main__":
    main()
